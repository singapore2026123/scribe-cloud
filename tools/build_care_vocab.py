# -*- coding: utf-8 -*-
"""Extract the care_watcher app's controlled care vocabulary (JA/EN/Burmese) from
多言語対応(データ).xlsx, convert the Burmese from Zawgyi->Unicode, and write
hf-space/care_vocab.json. Also grow burmese_lm_corpus.txt with the converted terms.

Language mapping (from M_MULTILINGUAL_MST): LANGUAGE1=English, 2=简体中文, 3=Tiếng Việt,
LANGUAGE4=Burmese(ミャンマー語). JA source = the NAME column (SST sheets) or LANGUAGE (M_MULTILINGUAL)."""
import openpyxl, json, os, sys
sys.path.insert(0, os.path.dirname(__file__))
from zawgyi import zg2uni

DRIVE = r"G:\My Drive\Singapore-PoC\poc-multilingual-voice"
SRC = os.path.join(DRIVE, "多言語対応(データ).xlsx")
HF = r"C:\Users\kanamic\Projects\voice-pipeline\scribe-cloud\hf-space"
OUT_VOCAB = os.path.join(HF, "care_vocab.json")
CORPUS = os.path.join(HF, "burmese_lm_corpus.txt")

# (sheet, ja_header, en_header, my_header)
SHEETS = [
    ("M_MULTILINGUAL", "LANGUAGE", "LANGUAGE1", "LANGUAGE4"),
    ("SST_M_CATEGORY", "NAME", "LANGUAGE1", "LANGUAGE4"),
    ("SST_M_ITEM",     "NAME", "LANGUAGE1", "LANGUAGE4"),
    ("SST_M_VALUE",    "NAME", "LANGUAGE1", "LANGUAGE4"),
]

def col(ws, name):
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value).strip() == name:
            return c
    return None

wb = openpyxl.load_workbook(SRC, data_only=True)
seen = {}
per_sheet = {}
for sheet, ja_h, en_h, my_h in SHEETS:
    ws = wb[sheet]
    cj, ce, cm = col(ws, ja_h), col(ws, en_h), col(ws, my_h)
    print(f"{sheet}: ja=col{cj} en=col{ce} my=col{cm}")
    if not cm:
        print("  !! no Burmese column found, skipping"); continue
    n = 0
    for r in range(2, ws.max_row + 1):
        my_z = ws.cell(r, cm).value
        if not my_z or not str(my_z).strip():
            continue
        my = zg2uni(str(my_z).strip())
        ja = str(ws.cell(r, cj).value).strip() if cj and ws.cell(r, cj).value else ""
        en = str(ws.cell(r, ce).value).strip() if ce and ws.cell(r, ce).value else ""
        key = my
        if key in seen:
            continue
        seen[key] = {"my": my, "ja": ja, "en": en, "src": sheet}
        n += 1
    per_sheet[sheet] = n
    print(f"  +{n} unique Burmese terms")

vocab = list(seen.values())
with open(OUT_VOCAB, "w", encoding="utf-8") as f:
    json.dump(vocab, f, ensure_ascii=False, indent=0)
print(f"\nwrote {len(vocab)} vocab entries -> {OUT_VOCAB}")
for v in vocab[:6]:
    print(f"  {v['ja']:<8} | {v['en'][:24]:<24} | {v['my']}")

# grow LM corpus (dedup against existing lines)
existing = set()
if os.path.exists(CORPUS):
    existing = set(l.strip() for l in open(CORPUS, encoding="utf-8") if l.strip())
new_lines = [v["my"] for v in vocab if v["my"] not in existing]
if new_lines:
    with open(CORPUS, "a", encoding="utf-8") as f:
        f.write("\n" + "\n".join(new_lines) + "\n")
print(f"corpus: +{len(new_lines)} new lines (was {len(existing)}) -> {len(existing)+len(new_lines)} total")
